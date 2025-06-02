import os
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import cell
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

load_dotenv()
FOLDER_PATH = os.getenv("FOLDER_PATH")

WRITE_START_ROW = 9 # adjust your excel sheets

def read_code(file_path):
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            read_line = f.readlines()
        return [line.rstrip('\n') for line in read_line]
    except FileNotFoundError:
        print("target code file not found")
        exit()

def write_code_with_pd(edit_file_name, sheet_name, read_file_name):
    target_path = f'{FOLDER_PATH}/{edit_file_name}'

    start_row = 10
    start_col = 2

    code_lines = read_code(read_file_name)
    df_code = pd.DataFrame(code_lines, columns=['code'])

    try:
        if os.path.exists(target_path):
            with pd.ExcelWriter(target_path, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
                if sheet_name not in writer.book.sheetnames:
                    writer.book.create_sheet(sheet_name)
                df_code.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, startcol=start_col)

        else:
            print("excel file not found")
    except Exception as e:
        print(e)
        exit()

def refresh_rows_in_print_area(file_name, sheet_name, read_file_name):
    target_path = f'{FOLDER_PATH}/{file_name}'

    code_lines = read_code(read_file_name)
    insert_num = len(code_lines)

    if not os.path.exists(target_path):
        print("target code file not found")
        exit()

    try:
        workbook = load_workbook(target_path)
        if sheet_name not in workbook.sheetnames:
            print("target sheet does not exist")
            exit()
        sheet = workbook[sheet_name]
        if not sheet.print_area:
            print("print area does not exist")
            exit()

        _, _, _, max_row = cell.range_boundaries(sheet.print_area)
        delete_lines = max_row - WRITE_START_ROW + 1
        sheet.delete_rows(idx=delete_lines, amount=delete_lines)
        print("deleted lines")

        sheet.insert_rows(idx=WRITE_START_ROW, amount=insert_num)
        print("inserted lines")

        workbook.save(target_path)
    except Exception as e:
        print(e)
        exit()

def refresh_rows_have_data(ws: Worksheet, read_file_name: str):
    code_lines = read_code(read_file_name)
    insert_num = len(code_lines)
    try:
        last_data_row = ws.max_row
        delete_lines = last_data_row - WRITE_START_ROW+ 1
        ws.delete_rows(idx=WRITE_START_ROW, amount=delete_lines)
        print(f"deleted {delete_lines} lines, until {last_data_row} row")

        ws.insert_rows(idx=WRITE_START_ROW, amount=insert_num)
        print("inserted lines")
    except Exception as e:
        print(e)
        exit()

def write_rows_with_pyxl(ws: Worksheet, read_file_name: str):
    code_lines = read_code(read_file_name)

    for idx, line in enumerate(code_lines):
        ws[f"D{WRITE_START_ROW+idx}"] = line
        font = Font(bold=False, name="Arial", size=10)
        ws[f"D{WRITE_START_ROW+idx}"].font = font

def apply_square_lattice(ws: Worksheet, start_row, end_row, start_col, end_col):
    lattice = Border(top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="thin"), left=Side(style="thin"))

    # left and right side border
    for row in range(start_row, end_row+1):
        ws.cell(row=row, column=start_col).border = Border(left=lattice.left)
        ws.cell(row=row, column=end_col).border = Border(right=lattice.right)

    # top and bottom side border
    for col in range(start_col, end_col+1):
        if col == start_col:
            ws.cell(row=start_row, column=col).border = Border(left=lattice.left, top=lattice.top)
            ws.cell(row=end_row, column=col).border = Border(left=lattice.left, bottom=lattice.top)
        elif col == end_col:
            ws.cell(row=start_row, column=col).border = Border(right=lattice.left, top=lattice.top)
            ws.cell(row=end_row, column=col).border = Border(right=lattice.left, bottom=lattice.top)
        else:
            ws.cell(row=start_row, column=col).border = Border(top=lattice.top)
            ws.cell(row=end_row, column=col).border = Border(bottom=lattice.bottom)

def apply_formatted_lattice(ws: Worksheet, row_num: int):
    lattice = Border(top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="thin"), left=Side(style="thin"))
    for num in range(row_num+1):
        row = num + WRITE_START_ROW
        if row == WRITE_START_ROW:
            ws.cell(row=row, column=2).border = Border(left=lattice.left, right=lattice.right)
            ws.cell(row=row, column=3).border = Border(left=lattice.left, top=lattice.top, right=lattice.right)
            ws.cell(row=row, column=12).border = Border(top=lattice.top, right=lattice.right)
            ws.cell(row=row, column=16).border = Border(top=lattice.top, right=lattice.right)
        elif row == row_num + WRITE_START_ROW:
            ws.cell(row=row, column=2).border = Border(left=lattice.left, right=lattice.right, bottom=lattice.bottom)
            ws.cell(row=row, column=3).border = Border(left=lattice.left, right=lattice.right, bottom=lattice.bottom)
            ws.cell(row=row, column=12).border = Border(right=lattice.right, bottom=lattice.bottom)
            ws.cell(row=row, column=16).border = Border(right=lattice.right, bottom=lattice.bottom)
        else:
            ws.cell(row=row, column=2).border = Border(left=lattice.left, right=lattice.right)
            ws.cell(row=row, column=3).border = Border(left=lattice.left, right=lattice.right)
            ws.cell(row=row, column=12).border = Border(right=lattice.right)
            ws.cell(row=row, column=16).border = Border(right=lattice.right)

    for row in ws.iter_rows(min_row=WRITE_START_ROW, max_row=row_num+WRITE_START_ROW, min_col=4, max_col=16):
        for cell in row:
            cell.border = Border(top=Side(style="thin"), bottom=Side(style="thin"))

    for row in range(row_num+1):
        ws.cell(row=row+WRITE_START_ROW, column=12).border = Border(right=lattice.right, top=lattice.top, bottom=lattice.bottom)
        ws.cell(row=row+WRITE_START_ROW, column=16).border = Border(right=lattice.right, top=lattice.top, bottom=lattice.bottom)

def change_sheet_name(ws: Worksheet, new_sheet_name: str) -> None:
    if len(new_sheet_name) > 31:
        new_sheet_name = new_sheet_name[:31]
    ws.title = new_sheet_name

def write_cell(ws: Worksheet, target_cell: str, data: int | str) -> None:
    ws[target_cell] = data

def fill_color(ws: Worksheet, start_row: int, end_row: int, color: str) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=2, max_col=2):
        for cell in row:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
